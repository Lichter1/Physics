"""
Build the report data-tables from the live workbook.

Single source of truth = the Excel workbook (same as every other script here).
This produces one nicely-formatted table per graph, *compressed* where several
graphs share the same underlying measurement block:

    Table 1  — Rotameter calibration                  -> Graph #1
    Table 2  — Venturi   (ΔP, Q_ideal, Q_real, Re, C)  -> Graphs #2, #6, #9
    Table 3  — Diaphragm (ΔP, Q_ideal, Q_real, Re, C)  -> Graphs #3, #7, #9
    Table 4  — Nozzle    (ΔP, Q_ideal, Q_real, Re, C)  -> Graphs #4, #8, #9
    Table 5  — Pitot velocities (Re, 3 methods, √ΔP)   -> Graph #5
    Table 6  — Rotameter pressure test                 -> Graph #10

So the ten graphs collapse into six data tables (graphs #2/#3/#4, #6/#7/#8 and
#9 all read the same per-instrument block, so each instrument is shown once with
every quantity its graphs need).  A leading "Fit summary" sheet collects the
fitted equations / slopes / χ²_ν for every graph.

Outputs (under figures/):
    tables.xlsx              — formatted, one sheet per table (+ Fit summary)
    tables/<name>.csv        — flat CSV (value & uncertainty columns) per table

Run:  python make_tables.py
"""

from __future__ import annotations

import csv
from pathlib import Path

import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

import flow_common as fc
import fig01_rotameter_calibration as fig01

# --- pull the same fit knobs the figure scripts use, to keep numbers identical
import fig10_rotameter_pressure as fig10

HERE = Path(__file__).resolve().parent
OUTDIR = HERE / "figures"
CSVDIR = OUTDIR / "tables"
XLSX = OUTDIR / "tables.xlsx"

INSTRUMENTS = [("Venturi", 2), ("Diaphragm", 3), ("Nozzle", 4)]


# --------------------------------------------------------------------------
# A tiny "table" description the writers below consume.
#   columns: list of dicts {group, name, unit, data (np array), fmt}
# --------------------------------------------------------------------------
def _col(group, name, unit, data, fmt):
    return dict(group=group, name=name, unit=unit, data=np.asarray(data, float), fmt=fmt)


def build_tables(wb):
    tables = []  # (key, sheet_title, caption, columns)

    # ---- Table 1 — Rotameter calibration (Graph #1) ----------------------
    ws = wb["Rotameter"]
    r = fc.read_column(ws, "E", 4, 13)            # reading [cm]
    q = fc.read_column(ws, "D", 4, 13)            # flow [L/min]
    q_err = fc.ldn_error(wb, q)                    # LDN 1009 datasheet error
    n = np.arange(1, len(r) + 1)
    tables.append((
        "T1_Rotameter", "Table 1 — Rotameter calibration",
        "Rotameter calibration at atmospheric pressure (Graph #1). "
        "Y-error is the LDN 1009 mass-flow-meter datasheet uncertainty; "
        "X-error (reading) = 0.05 cm.",
        [
            _col("", "Point", "", n, "0"),
            _col("", "Rotameter reading", "cm", r, "0.0"),
            _col("Flow rate Q", "Q", "L/min", q, "0.0"),
            _col("Flow rate Q", "ΔQ", "L/min", q_err, "0.00"),
        ],
    ))

    # ---- Tables 2-4 — per-instrument block (Graphs #2/6/9 etc.) ----------
    for sheet, tn in INSTRUMENTS:
        ws = wb[sheet]
        dP   = fc.read_column(ws, "E", 4, 13)
        dPe  = fc.read_column(ws, "F", 4, 13)
        Qi   = fc.read_column(ws, "G", 4, 13)
        Qie  = fc.read_column(ws, "H", 4, 13)
        Qr   = fc.read_column(ws, "I", 4, 13)
        Qre  = fc.read_column(ws, "J", 4, 13)
        Re   = fc.read_column(ws, "K", 4, 13)
        Ree  = fc.read_column(ws, "L", 4, 13)
        C    = fc.read_column(ws, "C", 19, 28)     # authoritative C-table
        with np.errstate(divide="ignore", invalid="ignore"):
            Ce = np.abs(C) * np.sqrt((Qre / Qr) ** 2 + (Qie / Qi) ** 2)
        n = np.arange(1, len(dP) + 1)
        graphs = {"Venturi": "#2, #6, #9", "Diaphragm": "#3, #7, #9",
                  "Nozzle": "#4, #8, #9"}[sheet]
        tables.append((
            f"T{tn}_{sheet}",
            f"Table {tn} — {sheet}",
            f"{sheet} measurement block (Graphs {graphs}). "
            f"C = Q_real/Q_ideal; ΔP-error is constant (= {np.nanmean(dPe):.3g} Pa).",
            [
                _col("", "Point", "", n, "0"),
                _col("Pressure ΔP", "ΔP", "Pa", dP, "0.0"),
                _col("Pressure ΔP", "Δ(ΔP)", "Pa", dPe, "0.00"),
                _col("Ideal flow Q_ideal", "Q_ideal", "L/min", Qi, "0.00"),
                _col("Ideal flow Q_ideal", "ΔQ_ideal", "L/min", Qie, "0.00"),
                _col("Real flow Q_real", "Q_real", "L/min", Qr, "0.00"),
                _col("Real flow Q_real", "ΔQ_real", "L/min", Qre, "0.00"),
                _col("Reynolds Re", "Re", "-", Re, "0"),
                _col("Reynolds Re", "ΔRe", "-", Ree, "0"),
                _col("Discharge coeff. C", "C", "-", C, "0.000"),
                _col("Discharge coeff. C", "ΔC", "-", Ce, "0.000"),
            ],
        ))

    # ---- Table 5 — Pitot velocities (Graph #5) --------------------------
    ws = wb["Pitot"]
    Re   = fc.read_column(ws, "N", 4, 13)
    Ree  = fc.read_column(ws, "O", 4, 13)
    u_r  = fc.read_column(ws, "L", 4, 13);  u_re = fc.read_column(ws, "M", 4, 13)
    u_p  = fc.read_column(ws, "I", 4, 13);  u_pe = fc.read_column(ws, "J", 4, 13)
    u_b  = fc.read_column(ws, "G", 4, 13);  u_be = fc.read_column(ws, "H", 4, 13)
    sP   = fc.read_column(ws, "E", 4, 13);  sPe  = fc.read_column(ws, "F", 4, 13)
    n = np.arange(1, len(Re) + 1)
    tables.append((
        "T5_Pitot", "Table 5 — Pitot velocities",
        "Pitot-tube velocities by three methods vs. Reynolds number and √ΔP "
        "(Graph #5).  Mirrors the reference report's Table 5.",
        [
            _col("", "Point", "", n, "0"),
            _col("Reynolds Re", "Re", "-", Re, "0"),
            _col("Reynolds Re", "ΔRe", "-", Ree, "0"),
            _col("Avg. velocity (rotameter)", "u", "m/s", u_r, "0.00"),
            _col("Avg. velocity (rotameter)", "Δu", "m/s", u_re, "0.00"),
            _col("Avg. velocity (Pitot, 1/7)", "u", "m/s", u_p, "0.00"),
            _col("Avg. velocity (Pitot, 1/7)", "Δu", "m/s", u_pe, "0.00"),
            _col("Velocity (Bernoulli, point)", "u", "m/s", u_b, "0.00"),
            _col("Velocity (Bernoulli, point)", "Δu", "m/s", u_be, "0.00"),
            _col("Pressure √ΔP", "√ΔP", "√Pa", sP, "0.00"),
            _col("Pressure √ΔP", "Δ√ΔP", "√Pa", sPe, "0.00"),
        ],
    ))

    # ---- Table 6 — Rotameter pressure test (Graph #10) ------------------
    ws = wb["Rotameter"]
    p_g  = fc.read_column(ws, "C", 17, 21)         # gauge pressure [bar]
    rd   = fc.read_column(ws, "D", 17, 21)         # reading [cm]
    qtru = fc.read_column(ws, "E", 17, 21)         # constant true flow [L/min]
    p_abs = p_g + fig10.ATM_BAR if fig10.PRESSURE_IS_GAUGE else p_g
    ratio = fig10.P_REF_BAR / p_abs                # ref_over_p
    x = np.sqrt(ratio)
    perr = fc.read_cell(wb["Setup"], "C15")
    x_err = 0.5 * x * (perr / p_abs) if np.isfinite(perr) else np.full_like(x, np.nan)
    cal = fig01.get_calibration(wb)
    q_calc = cal["slope"] * rd + cal["intercept"]
    q_calc_err = abs(cal["slope"]) * 0.05          # READING_ERR = 0.05 cm
    q_calc_err = np.full_like(q_calc, q_calc_err)
    n = np.arange(1, len(p_g) + 1)
    tables.append((
        "T6_Rotameter_Pressure", "Table 6 — Rotameter pressure test",
        f"Variable-pressure rotameter test at constant true flow "
        f"({np.nanmean(qtru):.0f} L/min) (Graph #10).  "
        f"P_abs = P_gauge + {fig10.ATM_BAR} bar; "
        f"Q_calc from the Graph #1 calibration (slope {cal['slope']:.3g}).",
        [
            _col("", "Point", "", n, "0"),
            _col("", "Gauge pressure", "bar", p_g, "0.00"),
            _col("", "Absolute pressure P_i", "bar_a", p_abs, "0.00"),
            _col("", "Rotameter reading", "cm", rd, "0.0"),
            _col("√(P_ref/P_i)", "x", "-", x, "0.0000"),
            _col("√(P_ref/P_i)", "Δx", "-", x_err, "0.0000"),
            _col("Calc. flow Q", "Q", "L/min", q_calc, "0.0"),
            _col("Calc. flow Q", "ΔQ", "L/min", q_calc_err, "0.0"),
        ],
    ))

    return tables


# --------------------------------------------------------------------------
# Fit summary (re-uses the figure scripts' own fits so numbers match exactly)
# --------------------------------------------------------------------------
def build_fit_summary(wb):
    rows = [("Graph", "Quantity", "Fitted relation", "Slope", "χ²_ν")]
    sqrt = np.sqrt

    cal = fig01.get_calibration(wb)
    rows.append(("#1", "Rotameter cal.",
                 f"Q = {cal['slope']:.4g}·h" + ("" if cal["through_origin"] else f" + {cal['intercept']:.3g}"),
                 f"{cal['slope']:.4g}", _chi_rotameter(wb, cal)))

    gmap = {"Venturi": "#2", "Diaphragm": "#3", "Nozzle": "#4"}
    for sheet, g in gmap.items():
        ws = wb[sheet]
        dP = fc.read_column(ws, "E", 4, 13); Qi = fc.read_column(ws, "G", 4, 13)
        Qr = fc.read_column(ws, "I", 4, 13)
        dPe = fc.read_column(ws, "F", 4, 13)
        Qie = fc.read_column(ws, "H", 4, 13); Qre = fc.read_column(ws, "J", 4, 13)
        m = fc.finite_mask(dP, Qi, Qr)
        ai, _ = fc.fit_sqrt(dP[m], Qi[m]); ar, _ = fc.fit_sqrt(dP[m], Qr[m])
        _, _, ci = fc.chi_square(dP[m], Qi[m], lambda t: ai * sqrt(t),
                                 dmodel=lambda t: ai / (2 * sqrt(t)),
                                 xerr=dPe[m], yerr=Qie[m], n_params=1)
        _, _, cr = fc.chi_square(dP[m], Qr[m], lambda t: ar * sqrt(t),
                                 dmodel=lambda t: ar / (2 * sqrt(t)),
                                 xerr=dPe[m], yerr=Qre[m], n_params=1)
        rows.append((g, f"{sheet} Q_ideal", f"Q = {ai:.4g}·√ΔP", f"{ai:.4g}", f"{ci:.3g}"))
        rows.append((g, f"{sheet} Q_real",  f"Q = {ar:.4g}·√ΔP", f"{ar:.4g}", f"{cr:.3g}"))

    # #5 Pitot
    ws = wb["Pitot"]
    sP = fc.read_column(ws, "E", 4, 13); sPe = fc.read_column(ws, "F", 4, 13)
    for col, ecol, name in [("L", "M", "u rotameter"), ("I", "J", "u Pitot 1/7"),
                            ("G", "H", "u Bernoulli")]:
        u = fc.read_column(ws, col, 4, 13); ue = fc.read_column(ws, ecol, 4, 13)
        m = fc.finite_mask(sP, u)
        a, _ = fc.fit_through_origin(sP[m], u[m])
        _, _, c = fc.chi_square(sP[m], u[m], lambda t: a * t,
                                dmodel=lambda t: np.full_like(np.asarray(t, float), a),
                                xerr=sPe[m], yerr=ue[m], n_params=1)
        rows.append(("#5", name, f"u = {a:.4g}·√ΔP", f"{a:.4g}", f"{c:.3g}"))

    # #6-8 C vs Re (no global fit) and #9 (slope = C, through origin)
    ws = wb["Summary_QvQ"]
    qvq = {"Venturi": ("C", "D", "#9 (Venturi)"), "Diaphragm": ("F", "G", "#9 (Diaphragm)"),
           "Nozzle": ("I", "J", "#9 (Nozzle)")}
    for src, (xc, yc, g) in qvq.items():
        x = fc.read_column(ws, xc, 4, 13); y = fc.read_column(ws, yc, 4, 13)
        xe = fc.read_column(wb[src], "H", 4, 13); ye = fc.read_column(wb[src], "J", 4, 13)
        m = fc.finite_mask(x, y)
        a, _ = fc.fit_through_origin(x[m], y[m])
        _, _, c = fc.chi_square(x[m], y[m], lambda t: a * t,
                                dmodel=lambda t: np.full_like(np.asarray(t, float), a),
                                xerr=xe[m], yerr=ye[m], n_params=1)
        rows.append((g, "Q_real vs Q_ideal", f"Q_real = {a:.4g}·Q_ideal  (slope = C)",
                     f"{a:.4g}", f"{c:.3g}"))

    # #10 rotameter pressure
    res = fig10.main(wb)
    if res.get("status") == "ok":
        b = res["intercept"]
        relation = f"Q = {res['slope']:.4g}·x {'+' if b >= 0 else '−'} {abs(b):.4g}"
        rows.append(("#10", "Rotameter pressure", relation,
                     f"{res['slope']:.4g}", f"{res['chi2_red']:.3g}"))
    return rows


def _chi_rotameter(wb, cal):
    yerr = fc.ldn_error(wb, cal["y"])
    n_params = 1 if cal["through_origin"] else 2
    _, _, c = fc.chi_square(cal["x"], cal["y"],
                            lambda t: cal["slope"] * t + cal["intercept"],
                            dmodel=lambda t: np.full_like(np.asarray(t, float), cal["slope"]),
                            xerr=0.05, yerr=yerr, n_params=n_params)
    return f"{c:.3g}"


# --------------------------------------------------------------------------
# Writers
# --------------------------------------------------------------------------
THIN = Side(style="thin", color="999999")
MED = Side(style="medium", color="333333")
HDR_FILL = PatternFill("solid", fgColor="DCE6F1")
GRP_FILL = PatternFill("solid", fgColor="C5D9F1")
TITLE_FONT = Font(bold=True, size=13)
GRP_FONT = Font(bold=True, size=11)
HDR_FONT = Font(bold=True, size=10)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)


def write_sheet(wb_out, key, title, caption, columns):
    ws = wb_out.create_sheet(title=key[:31])
    ncol = len(columns)
    # Row 1: title
    ws.cell(1, 1, title).font = TITLE_FONT
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncol)
    # Row 2: caption
    cap = ws.cell(2, 1, caption)
    cap.font = Font(italic=True, size=9, color="555555")
    cap.alignment = Alignment(wrap_text=True, vertical="top")
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=ncol)
    ws.row_dimensions[2].height = 28

    grp_row, hdr_row, unit_row, first_data = 3, 4, 5, 6

    # Group header (row 3) — merge consecutive equal group labels
    c = 1
    while c <= ncol:
        g = columns[c - 1]["group"]
        span = 1
        while c + span <= ncol and columns[c + span - 1]["group"] == g and g != "":
            span += 1
        if g:
            cell = ws.cell(grp_row, c, g)
            cell.font, cell.fill, cell.alignment = GRP_FONT, GRP_FILL, CENTER
            if span > 1:
                ws.merge_cells(start_row=grp_row, start_column=c,
                               end_row=grp_row, end_column=c + span - 1)
        c += span

    # Symbol (row 4) + unit (row 5)
    for j, col in enumerate(columns, start=1):
        h = ws.cell(hdr_row, j, col["name"]); h.font, h.fill, h.alignment = HDR_FONT, HDR_FILL, CENTER
        u = ws.cell(unit_row, j, f"[{col['unit']}]" if col["unit"] else "")
        u.font = Font(size=9, italic=True); u.alignment = CENTER; u.fill = HDR_FILL

    # Data
    nrows = len(columns[0]["data"])
    for i in range(nrows):
        for j, col in enumerate(columns, start=1):
            v = col["data"][i]
            cell = ws.cell(first_data + i, j)
            if np.isfinite(v):
                cell.value = float(v)
                cell.number_format = col["fmt"]
            else:
                cell.value = "—"
            cell.alignment = CENTER

    # Borders + widths
    last = first_data + nrows - 1
    for r in range(grp_row, last + 1):
        for j in range(1, ncol + 1):
            ws.cell(r, j).border = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
    for j in range(1, ncol + 1):
        ws.cell(unit_row, j).border = Border(left=THIN, right=THIN, top=THIN, bottom=MED)
    for j, col in enumerate(columns, start=1):
        w = max(len(col["name"]), len(col["unit"]) + 2, 7) + 2
        ws.column_dimensions[get_column_letter(j)].width = w
    ws.freeze_panes = ws.cell(first_data, 1)


def write_summary(wb_out, rows):
    ws = wb_out.create_sheet(title="Fit summary")
    ws.cell(1, 1, "Fit summary — all graphs").font = TITLE_FONT
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(rows[0]))
    for j, h in enumerate(rows[0], start=1):
        c = ws.cell(3, j, h); c.font, c.fill, c.alignment = HDR_FONT, HDR_FILL, CENTER
    for i, row in enumerate(rows[1:], start=4):
        for j, v in enumerate(row, start=1):
            c = ws.cell(i, j, v)
            c.alignment = Alignment(horizontal="center" if j != 3 else "left", vertical="center")
            c.border = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
    widths = [14, 22, 34, 12, 10]
    for j, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(j)].width = w
    ws.freeze_panes = "A4"


def write_csv(key, title, caption, columns):
    CSVDIR.mkdir(parents=True, exist_ok=True)
    path = CSVDIR / f"{key}.csv"
    with open(path, "w", newline="", encoding="utf-8-sig") as f:
        w = csv.writer(f)
        w.writerow([title])
        w.writerow([caption])
        w.writerow([f"{c['name']} [{c['unit']}]" if c["unit"] else c["name"] for c in columns])
        nrows = len(columns[0]["data"])
        for i in range(nrows):
            w.writerow([(round(float(c["data"][i]), 6) if np.isfinite(c["data"][i]) else "")
                        for c in columns])
    return path


def main():
    wb = fc.load_workbook()
    tables = build_tables(wb)
    summary = build_fit_summary(wb)

    out = Workbook()
    out.remove(out.active)  # drop default sheet
    write_summary(out, summary)
    for key, title, caption, columns in tables:
        write_sheet(out, key, title, caption, columns)
        write_csv(key, title, caption, columns)
    OUTDIR.mkdir(parents=True, exist_ok=True)
    out.save(XLSX)
    print(f"Wrote {XLSX.relative_to(HERE)}  ({len(tables)} tables + Fit summary)")
    print(f"CSVs in {CSVDIR.relative_to(HERE)}/ :")
    for key, *_ in tables:
        print(f"   {key}.csv")


if __name__ == "__main__":
    main()
