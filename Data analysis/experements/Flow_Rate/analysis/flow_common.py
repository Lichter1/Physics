"""
Shared utilities for the "Flow Rate and Velocity Measurement" lab figures.

Everything the individual ``fig*.py`` scripts have in common lives here:
  * locating and loading the Excel workbook (CALCULATED values only),
  * reading a numeric column into a NumPy array (blank cells -> NaN),
  * a few small curve fits (linear, linear-through-origin, square-root),
  * consistent matplotlib styling + figure saving (PNG @ 200 dpi + PDF),
  * a "zoom inset" helper and an "awaiting data" placeholder image.

You normally do NOT need to edit this file to tweak a single graph -- the
per-figure knobs live in the CONFIG block at the top of each ``fig*.py``.
Edit here only to change something GLOBAL: the workbook file name, the output
folder, or the overall plotting style.

Design note
-----------
The workbook is the single source of truth.  Every "_err" and every derived
quantity (ΔP, Q_ideal, Re, ...) is read straight from the green/calculated
cells -- we never re-derive physics here.  The only quantities computed in
Python are the trendlines the report asks for (slopes/fits) and, for the
C-vs-Re graphs, the error bar on C (the workbook has no C_err column), which
is obtained by standard error propagation from the Q_real / Q_ideal errors.
"""

from __future__ import annotations

import os
from pathlib import Path

import matplotlib

matplotlib.use("Agg")  # headless: write files, never open a window
import matplotlib.pyplot as plt  # noqa: E402
import numpy as np  # noqa: E402

try:
    import openpyxl  # noqa: E402
except ImportError as exc:  # pragma: no cover
    raise SystemExit(
        "openpyxl is required.  Install the dependencies with:\n"
        "    pip install -r requirements.txt\n"
        f"(original error: {exc})"
    )

# --------------------------------------------------------------------------
# Paths  (override the workbook with:  FLOW_WORKBOOK=/path/to.xlsx python ...)
# --------------------------------------------------------------------------
HERE = Path(__file__).resolve().parent
DATA_DIR = HERE.parent  # the Flow_Rate folder, where the workbook lives


def _discover_workbook() -> Path:
    """Locate the lab workbook.

    Priority:
      1. the FLOW_WORKBOOK environment variable, if set;
      2. the single .xlsx in the parent (Flow_Rate) folder.
    Auto-discovery is used because the Hebrew file name has been renamed before
    (spaces vs. underscores, "copy" vs. "1") and Unicode normalisation of the
    name is unreliable across systems.  If several .xlsx files are present we
    prefer one whose name contains "ספיק"/"מהיר" (flow/velocity).
    """
    env = os.environ.get("FLOW_WORKBOOK")
    if env:
        return Path(env)
    candidates = [p for p in DATA_DIR.glob("*.xlsx")
                  if not p.name.startswith("~$")
                  and "BACKUP" not in p.name.upper()]
    if not candidates:
        # Return the conventional name so the error message is informative.
        return DATA_DIR / "מדידת ספיקה ומהירות לוח מדידה.xlsx"
    if len(candidates) > 1:
        preferred = [p for p in candidates
                     if ("ספיק" in p.name or "מהיר" in p.name)]
        if preferred:
            candidates = preferred
    return sorted(candidates)[0]


WORKBOOK = _discover_workbook()
FIGDIR = Path(os.environ.get("FLOW_FIGDIR", HERE / "figures"))

# Spreadsheet geometry (1-based, inclusive) -- shared by several scripts.
INSTRUMENT_ROWS = (4, 13)        # the 10 measurement points
CTABLE_ROWS = (19, 28)           # the C = Qreal/Qideal block below the table
PITOT_ROWS = (4, 13)             # the 10 Pitot points
ROTAMETER_CAL_ROWS = (4, 13)     # part 1: atmospheric calibration (10 points)
ROTAMETER_PRES_ROWS = (17, 21)   # part 2: variable absolute pressure (5 points)


# --------------------------------------------------------------------------
# Workbook reading
# --------------------------------------------------------------------------
def _formula_values(path: str) -> dict:
    """Evaluate every Excel formula in the workbook and return {(SHEET, CELL): value}.

    Used only as a fallback for formula cells whose cached value is missing
    (e.g. right after openpyxl re-saved the file and dropped Excel's cache).
    The ``formulas`` engine evaluates the workbook's OWN formulas, so we are
    still reading "what Excel computes", never a Python reimplementation.

    ``formulas`` chokes on non-ASCII / spaced file names, so we evaluate a
    temporary ASCII copy.
    """
    import shutil
    import tempfile

    try:
        import formulas  # noqa: F401
    except ImportError:
        return {}  # engine unavailable -> rely on whatever cache exists

    tmpdir = tempfile.mkdtemp(prefix="flow_recalc_")
    try:
        ascii_path = os.path.join(tmpdir, "wb.xlsx")
        shutil.copy(path, ascii_path)
        xl = formulas.ExcelModel().loads(ascii_path).finish()
        sol = xl.calculate()
        # Keys look like  "'[wb.xlsx]SHEETNAME'!CELL"  (sheet upper-cased by the
        # engine, filename kept as-is).  Only one workbook is loaded, so parse
        # structurally rather than matching a case-sensitive filename prefix.
        out = {}
        for key, cell in sol.items():
            if not isinstance(key, str) or "]" not in key or "'!" not in key:
                continue
            sheet = key.split("]", 1)[1].split("'!", 1)[0]
            coord = key.rsplit("'!", 1)[1]
            if ":" in coord:  # skip range keys
                continue
            out[(sheet.upper(), coord.upper())] = _scalar(cell)
        return out
    except Exception:
        return {}
    finally:
        shutil.rmtree(tmpdir, ignore_errors=True)


def _scalar(cell):
    """Extract a plain Python scalar from a ``formulas`` result object."""
    val = getattr(cell, "value", cell)
    if isinstance(val, np.ndarray):
        if val.size == 0:
            return None
        val = val.ravel()[0]
    if isinstance(val, np.generic):
        try:
            val = val.item()
        except Exception:
            pass
    return val


class _MergedCell:
    __slots__ = ("value", "coordinate")

    def __init__(self, value, coordinate):
        self.value = value
        self.coordinate = coordinate


class _MergedSheet:
    """openpyxl-like sheet whose cell values prefer the cache, then recompute."""

    def __init__(self, book, name):
        self._b = book
        self.title = name

    def _value(self, coord):
        litv = self._b._lit[self.title][coord].value
        if litv is not None:
            return litv  # literal input, or a valid cached formula result
        fmlv = self._b._fml[self.title][coord].value
        if isinstance(fmlv, str) and fmlv.startswith("="):
            return self._b._fvals().get((self.title.upper(), coord.upper()))
        return litv

    def __getitem__(self, coord):
        return _MergedCell(self._value(coord), coord)

    def iter_rows(self, **kw):
        for row in self._b._fml[self.title].iter_rows(**kw):
            yield tuple(_MergedCell(self._value(c.coordinate), c.coordinate)
                        for c in row)

    @property
    def dimensions(self):
        return self._b._fml[self.title].dimensions

    @property
    def max_row(self):
        return self._b._fml[self.title].max_row

    @property
    def max_column(self):
        return self._b._fml[self.title].max_column


class MergedBook:
    """Read-only workbook view: cached values where present, recomputed otherwise.

    Behaves enough like an openpyxl workbook for our readers and the integrity
    check (``wb[sheet][cell].value``, ``wb.worksheets``, ``ws.iter_rows()``).
    The ``formulas`` recompute is lazy: if Excel's cache is intact it is never
    invoked.
    """

    def __init__(self, path):
        self._path = str(path)
        self._lit = openpyxl.load_workbook(path, data_only=True)
        self._fml = openpyxl.load_workbook(path, data_only=False)
        self._cache = None

    def _fvals(self):
        if self._cache is None:
            self._cache = _formula_values(self._path)
        return self._cache

    @property
    def sheetnames(self):
        return self._fml.sheetnames

    def __getitem__(self, name):
        return _MergedSheet(self, name)

    @property
    def worksheets(self):
        return [_MergedSheet(self, n) for n in self.sheetnames]


def load_workbook(path: Path | None = None):
    """Load the workbook for reading CALCULATED values.

    Returns a :class:`MergedBook` that yields Excel's cached value for every
    cell, and -- only when a formula's cache is missing -- transparently
    recomputes it from the workbook's own formulas via the ``formulas`` engine.
    This keeps the figures correct even immediately after the workbook has been
    edited programmatically (which drops Excel's cached values).
    """
    path = Path(path or WORKBOOK)
    if not path.exists():
        raise FileNotFoundError(f"Workbook not found: {path}")
    return MergedBook(path)


def _to_float(value):
    """Convert a cell value to float, mapping blanks / text / errors to NaN."""
    if value is None:
        return np.nan
    if isinstance(value, str):
        s = value.strip().replace(",", ".")
        if s == "":
            return np.nan
        try:
            return float(s)
        except ValueError:
            return np.nan  # e.g. a leftover "#DIV/0!" or a header
    try:
        return float(value)
    except (TypeError, ValueError):
        return np.nan


def read_column(ws, col: str, r0: int, r1: int) -> np.ndarray:
    """Read ``col`` over rows ``r0..r1`` (inclusive) into a float array.

    Blank cells, text and Excel errors become ``np.nan`` so downstream code
    can simply mask them out with :func:`finite_mask`.
    """
    return np.array([_to_float(ws[f"{col}{r}"].value) for r in range(r0, r1 + 1)],
                    dtype=float)


def read_cell(ws, addr: str):
    """Read a single cell as float (NaN if blank/non-numeric)."""
    return _to_float(ws[addr].value)


def finite_mask(*arrays: np.ndarray) -> np.ndarray:
    """Boolean mask of rows where *every* supplied array is finite."""
    mask = np.ones(len(arrays[0]), dtype=bool)
    for a in arrays:
        mask &= np.isfinite(a)
    return mask


def has_data(*arrays: np.ndarray) -> bool:
    """True if there is at least one row where all arrays are finite."""
    return bool(np.any(finite_mask(*arrays)))


def ldn_error(wb, Q):
    """LDN 1009 GAPL datasheet flow uncertainty (quadrature), from Setup C33:C36.

        ΔQ = sqrt( (pMV·Q)^2 + (pFS·FS)^2 + (pR·Q)^2 )

    Falls back to the legacy constant Setup!C14 if the spec block is absent.
    """
    s = wb["Setup"]
    pMV, pFS, FS, pR = (read_cell(s, "C33"), read_cell(s, "C34"),
                        read_cell(s, "C35"), read_cell(s, "C36"))
    Q = np.asarray(Q, float)
    if not all(np.isfinite(v) for v in (pMV, pFS, FS, pR)):
        return np.full_like(Q, read_cell(s, "C14"))
    return np.sqrt((pMV * Q) ** 2 + (pFS * FS) ** 2 + (pR * Q) ** 2)


# --------------------------------------------------------------------------
# Curve fits  (the only physics-free maths we do ourselves)
# --------------------------------------------------------------------------
def _r_squared(y, y_pred) -> float:
    y = np.asarray(y, float)
    ss_res = float(np.sum((y - y_pred) ** 2))
    ss_tot = float(np.sum((y - np.mean(y)) ** 2))
    return 1.0 - ss_res / ss_tot if ss_tot > 0 else float("nan")


def fit_linear(x, y):
    """Ordinary least squares  y = a*x + b.  Returns (a, b, r2)."""
    x, y = np.asarray(x, float), np.asarray(y, float)
    a, b = np.polyfit(x, y, 1)
    return a, b, _r_squared(y, a * x + b)


def fit_through_origin(x, y):
    """Least squares  y = a*x  (forced through the origin).  Returns (a, r2)."""
    x, y = np.asarray(x, float), np.asarray(y, float)
    a = float(np.sum(x * y) / np.sum(x * x))
    return a, _r_squared(y, a * x)


def fit_sqrt(x, y):
    """Least squares  y = a*sqrt(x).  Returns (a, r2).

    Used for the 'actual flow rate' calibration curves where physics expects
    Q ∝ √ΔP rather than a straight line.
    """
    x, y = np.asarray(x, float), np.asarray(y, float)
    sx = np.sqrt(x)
    a = float(np.sum(sx * y) / np.sum(sx * sx))
    return a, _r_squared(y, a * sx)


def smooth_x(x, n: int = 200):
    """A dense, sorted x-grid from 0 to max(x) for drawing smooth trend curves."""
    x = np.asarray(x, float)
    return np.linspace(0.0, float(np.nanmax(x)), n)


def chi_square(x, y, model, *, dmodel=None, xerr=None, yerr=None, n_params=2):
    """Goodness-of-fit chi-square with the effective-variance method.

        χ² = Σ (y_i − f(x_i))² / σ_i² ,   σ_i² = σ_y,i² + (f'(x_i)·σ_x,i)²

    ``model`` is the fitted f(x); ``dmodel`` is f'(x) (used to fold the x-error
    into σ via the local slope).  Returns (χ², ν, χ²/ν) where ν = N − n_params.
    """
    x = np.asarray(x, float)
    y = np.asarray(y, float)
    resid = y - model(x)
    sx = np.zeros_like(x) if xerr is None else np.nan_to_num(np.asarray(xerr, float))
    sy = np.zeros_like(x) if yerr is None else np.nan_to_num(np.asarray(yerr, float))
    var = sy ** 2
    if dmodel is not None:
        var = var + (np.asarray(dmodel(x), float) * sx) ** 2
    good = var > 0
    if not good.any():
        return float("nan"), 0, float("nan")
    chi2 = float(np.sum(resid[good] ** 2 / var[good]))
    dof = max(int(good.sum()) - n_params, 1)
    return chi2, dof, chi2 / dof


def fmt_chi2(chi2_red) -> str:
    """Legend snippet for a reduced chi-square value."""
    if chi2_red is None or not np.isfinite(chi2_red):
        return ""
    return rf"$\chi^2_\nu$ = {chi2_red:.2f}"


# --------------------------------------------------------------------------
# Plotting style + saving
# --------------------------------------------------------------------------
# A small, distinguishable palette (color-blind friendly-ish).
COLORS = {
    "ideal": "#1f77b4",     # blue
    "real": "#d62728",      # red
    "bernoulli": "#1f77b4",
    "profile17": "#2ca02c",  # green
    "rotameter": "#d62728",
    "venturi": "#1f77b4",
    "diaphragm": "#d62728",
    "nozzle": "#2ca02c",
    "literature": "#7f7f7f",  # grey
    "fit": "#000000",
}


def apply_style():
    """Apply a consistent, submission-quality matplotlib style."""
    plt.rcParams.update({
        "figure.figsize": (8.0, 5.5),
        "figure.dpi": 110,
        "savefig.dpi": 200,
        "font.size": 12,
        "axes.titlesize": 13,
        "axes.labelsize": 12,
        "axes.grid": True,
        "grid.alpha": 0.35,
        "grid.linestyle": "--",
        "legend.fontsize": 10,
        "legend.framealpha": 0.9,
        "errorbar.capsize": 3,
        "lines.markersize": 6,
        "axes.axisbelow": True,
    })


def errorbar_series(ax, x, y, *, xerr=None, yerr=None, color, label,
                    fmt="o", markersize=6, alpha=1.0, zorder=3):
    """A standard measured-point series with error bars on both axes."""
    return ax.errorbar(
        x, y, xerr=xerr, yerr=yerr, fmt=fmt, color=color, label=label,
        markersize=markersize, markerfacecolor=color, markeredgecolor="black",
        markeredgewidth=0.5, ecolor=color, elinewidth=1.0, capsize=3,
        alpha=alpha, linestyle="none", zorder=zorder,
    )


def add_zoom_inset(ax, draw_fn, xlim, ylim=None,
                   bounds=(0.56, 0.10, 0.40, 0.40), label="zoom"):
    """Create a zoomed inset and let ``draw_fn(axins)`` populate it.

    ``draw_fn`` receives the inset Axes and should re-plot the same series
    (error bars + trend lines) so the tiny X-errors become visible, exactly
    as the lab manual requires (cf. reference report figs. 7 & 12).
    """
    axins = ax.inset_axes(bounds)
    draw_fn(axins)
    axins.set_xlim(*xlim)
    if ylim is not None:
        axins.set_ylim(*ylim)
    axins.grid(True, alpha=0.35, linestyle="--")
    axins.tick_params(labelsize=8)
    ax.indicate_inset_zoom(axins, edgecolor="black", alpha=0.6)
    return axins


def save_fig(fig, name: str, formats=("png", "pdf")):
    """Save ``fig`` to the figures folder as ``name.<ext>`` for each format."""
    FIGDIR.mkdir(parents=True, exist_ok=True)
    paths = []
    for ext in formats:
        p = FIGDIR / f"{name}.{ext}"
        fig.savefig(p, bbox_inches="tight")
        paths.append(p)
    plt.close(fig)
    return paths


def placeholder(name: str, title: str,
                message="Awaiting measurements — fill the yellow input\n"
                        "cells in the Excel workbook, then re-run this script."):
    """Write a clearly-labelled 'no data yet' figure so the pipeline is visible.

    This is produced whenever the relevant input cells are still empty.  Once
    real data is entered, the same script draws the real graph instead.
    """
    apply_style()
    fig, ax = plt.subplots()
    ax.set_title(title)
    ax.text(0.5, 0.5, message, ha="center", va="center", fontsize=12,
            color="#b00020", transform=ax.transAxes,
            bbox=dict(boxstyle="round,pad=0.6", fc="#fff3f3", ec="#b00020"))
    ax.set_xticks([])
    ax.set_yticks([])
    ax.grid(False)
    return save_fig(fig, name)


def fmt_eq(slope, intercept=None, *, xsym="x", ysym="y", unit="", chi2_red=None):
    """Human-readable fit equation for legends/annotations."""
    if intercept is None:
        s = f"{ysym} = {slope:.4g}·{xsym}"
    else:
        sign = "+" if intercept >= 0 else "−"
        s = f"{ysym} = {slope:.4g}·{xsym} {sign} {abs(intercept):.4g}"
    if unit:
        s += f"  [{unit}]"
    if chi2_red is not None and np.isfinite(chi2_red):
        s += "\n" + fmt_chi2(chi2_red)
    return s
